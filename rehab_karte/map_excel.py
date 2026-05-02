import openpyxl
import os

path = "templates/master_template.xlsx"

def map_fields(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    mapping = {}
    
    # Search the first 100x30 area for keywords
    for r in range(1, 101):
        for c in range(1, 31):
            cell = ws.cell(r, c)
            val = str(cell.value) if cell.value else ""
            
            # Simple keyword matching
            if "氏名" in val and len(val) < 10:
                mapping["patient_name"] = (r, c+1) # Assuming adjacent
            elif "病名" in val:
                mapping["disease_name"] = (r, c+1)
            elif "発症日" in val:
                mapping["onset_date"] = (r, c+1)
            elif "手術日" in val:
                mapping["surgery_date"] = (r, c+1)
            elif "食事" in val and r > 30: # Likely ADL section
                mapping["adl_eating"] = (r, c) # Root
            elif "短期目標" in val:
                mapping["goal_short"] = (r+1, c)
            elif "長期目標" in val:
                mapping["goal_long"] = (r+1, c)

    print("--- Detected Mapping (Heuristic) ---")
    for k, v in mapping.items():
        addr = ws.cell(v[0], v[1]).coordinate
        print(f"{k}: {addr}")

if __name__ == "__main__":
    map_fields(path)
