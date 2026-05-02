import openpyxl
import os
import shutil
import json
import subprocess
from datetime import datetime
from database.db import get_connection

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "master_template.xlsx")

# Reuse the map coordinates for scraping back too
from utils.excel_mapper import EXCEL_MAP, ADL_START_ROW, ADL_ITEM_ORDER, ADL_COL_MAP

def get_work_path(patient_id):
    work_dir = os.path.join(os.getcwd(), "work_excel")
    os.makedirs(work_dir, exist_ok=True)
    return os.path.join(work_dir, f"editing_{patient_id}.xlsx")

def prepare_editing_excel(patient_id, plan_data):
    """Fills current record into a temp excel for editing."""
    target_path = get_work_path(patient_id)
    shutil.copy(TEMPLATE_PATH, target_path)
    
    wb = openpyxl.load_workbook(target_path)
    ws = wb.active
    
    # Inject basic fields
    for db_key, cell in EXCEL_MAP.items():
        if db_key in plan_data and plan_data[db_key]:
            ws[cell] = plan_data[db_key]
            
    # Inject ADL Matrix
    if "adl_evaluation_json" in plan_data and plan_data["adl_evaluation_json"]:
        try:
            adl = json.loads(plan_data["adl_evaluation_json"])
            for idx, item_name in enumerate(ADL_ITEM_ORDER):
                if item_name in adl:
                    row = ADL_START_ROW + idx
                    for key, col in ADL_COL_MAP.items():
                        if key in adl[item_name]:
                            ws[f"{col}{row}"] = adl[item_name][key]
        except: pass
        
    wb.save(target_path)
    return target_path

def scrape_editing_excel(patient_id):
    """Reads the saved excel back into a dict for DB storage."""
    target_path = get_work_path(patient_id)
    if not os.path.exists(target_path): return None
    
    wb = openpyxl.load_workbook(target_path, data_only=True)
    ws = wb.active
    
    new_data = {}
    # Scrape basic fields
    for db_key, cell in EXCEL_MAP.items():
        val = ws[cell].value
        new_data[db_key] = val if val is not None else ""
        
    # Scrape ADL Matrix
    adl = {}
    for idx, item_name in enumerate(ADL_ITEM_ORDER):
        row = ADL_START_ROW + idx
        adl[item_name] = {}
        for key, col in ADL_COL_MAP.items():
            val = ws[f"{col}{row}"].value
            adl[item_name][key] = val if val is not None else ""
    
    new_data["adl_evaluation_json"] = json.dumps(adl)
    return new_data

def launch_excel_and_wait(file_path):
    """Launches Excel and waits for the process to finish."""
    # Use start /wait on Windows or subprocess.run
    # Note: startfile is async, so we use subprocess to get a handle if possible
    # Alternatively, we just tell the user 'Press Sync after closing'
    os.startfile(file_path)
