import json
import openpyxl

def generate_hard_locked_html(dna_path, output_path):
    with open(dna_path, "r", encoding="utf-8") as f:
        dna = json.load(f)
    
    # Constants for conversion
    W_SCALE = 8.0  # px per Excel width unit
    H_SCALE = 1.33 # px per Excel point (standard)
    
    # 1. Pre-calculate X and Y offsets for columns and rows
    col_letters = sorted(dna["cols"].keys(), key=lambda x: openpyxl.utils.column_index_from_string(x))
    col_offsets = {}
    current_x = 0
    for letter in col_letters:
        col_offsets[letter] = current_x
        current_x += dna["cols"].get(letter, 8.38) * W_SCALE
    
    row_offsets = {}
    current_y = 0
    for r in range(1, 121):
        row_offsets[r] = current_y
        current_y += dna["rows"].get(str(r), 15.0) * H_SCALE

    # 2. Map Merged Ranges to Absolute Boxes
    merges = []
    skip_cells = set()
    for m_range in dna["merges"]:
        r_start, c_start, r_end, c_end = openpyxl.utils.range_boundaries(m_range)
        
        # Calculate Box (Top-Left to Bottom-Right)
        x = col_offsets[openpyxl.utils.get_column_letter(c_start)]
        y = row_offsets[r_start]
        
        # Width: Sum of widths of columns in range
        w = 0
        for c in range(c_start, c_end + 1):
            w += dna["cols"].get(openpyxl.utils.get_column_letter(c), 8.38) * W_SCALE
            
        # Height: Sum of heights of rows in range
        h = 0
        for r in range(r_start, r_end + 1):
            h += dna["rows"].get(str(r), 15.0) * H_SCALE
            
        merges.append({
            "r1": r_start, "c1": c_start, "r2": r_end, "c2": c_end,
            "x": x, "y": y, "w": w, "h": h
        })
        
        # Mark cells to skip in individual processing
        for r in range(r_start, r_end + 1):
            for c in range(c_start, c_end + 1):
                skip_cells.add((r, c))

    # 3. Categorize Cells
    cell_values = {(item["r"], item["c"]): item["v"] for item in dna["cells"]}

    # 4. Generate HTML
    html = ["""
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <title>統合臨床エディタ - ハードロック再現版</title>
    <style>
        body { background: #2c2c2c; font-family: "Yu Gothic UI", sans-serif; margin: 0; padding: 20px; overflow-x: auto; }
        .toolbar { background: #111; color: #fff; width: 100%; max-width: 1700px; padding: 15px; display: flex; gap: 20px; align-items: center; position: sticky; top: 0; z-index: 1000; box-shadow: 0 4px 20px rgba(0,0,0,0.5); margin-bottom: 20px; }
        .btn { padding: 12px 28px; cursor: pointer; border: none; font-weight: bold; border-radius: 4px; font-size: 11pt; transition: 0.2s; background: #28a745; color: white; }
        
        /* The Canvas: Immutable Structure */
        .spread-container { position: relative; background: #fff; box-shadow: 0 0 50px rgba(0,0,0,0.8); margin: 0 auto; border: 1px solid #999; }
        
        .box { position: absolute; box-sizing: border-box; border: 1px solid #000; font-size: 8pt; display: flex; align-items: center; overflow: hidden; }
        .label { background: #f8f8f8; font-weight: bold; padding: 1px; justify-content: center; text-align: center; }
        .input-box { background: #fff; }
        
        input, textarea { border: none; width: 100%; height: 100%; background: transparent; font-family: inherit; font-size: inherit; outline: none; padding: 1px; margin: 0; display: block; }
        textarea { resize: none; overflow: hidden; }
        input:focus, textarea:focus { background: #f0f7ff; box-shadow: inset 0 0 5px rgba(0,100,255,0.2); }
        
        @media print {
            .toolbar { display: none; }
            body { background: white; padding: 0; }
            .spread-container { box-shadow: none; border: none; left: 0 !important; top: 0 !important; }
        }
    </style>
</head>
<body>
    <div class="toolbar">
        <div style="font-size: 16pt; font-weight: bold;">様式第21号 ハードロック・クローン <span style="font-size: 8pt; opacity: 0.6;">(39Cols / 217Merges)</span></div>
        <button class="btn" onclick="saveData()">保存(Sync DB)</button>
        <button class="btn" style="background:#555;" onclick="window.close()">閉じる</button>
    </div>
    """]

    # Calculate Total Canvas Dimensions
    canvas_w = current_x + 50
    canvas_h = current_y + 100
    html.append(f'<div class="spread-container" style="width: {canvas_w}px; height: {canvas_h}px;">')

    # 5. Render Merged Boxes First
    for m in merges:
        val = ""
        # Check if any cell in range has value
        for r in range(m["r1"], m["r2"]+1):
            for c in range(m["c1"], m["c2"]+1):
                if (r, c) in cell_values:
                    val = cell_values[(r, c)]
                    break
            if val: break
            
        is_label = val != ""
        klass = "label" if is_label else "input-box"
        
        # ID for storage: using top-left
        col_letter = openpyxl.utils.get_column_letter(m["c1"])
        field_id = f"cell_{col_letter}{m['r1']}"
        
        content = val if is_label else f'<input type="text" id="{field_id}">'
        if not is_label and m["h"] > 30: # Multi-line if box is tall
             content = f'<textarea id="{field_id}"></textarea>'
             
        html.append(f'  <div class="box {klass}" style="left: {m["x"]}px; top: {m["y"]}px; width: {m["w"]}px; height: {m["h"]}px;">{content}</div>')

    # 6. Render Individual Non-Merged Cells
    for r in range(1, 121):
        for c_idx, letter in enumerate(col_letters):
            c = c_idx + 1
            if (r, c) in skip_cells:
                continue
            
            x = col_offsets[letter]
            y = row_offsets[r]
            w = dna["cols"].get(letter, 8.38) * W_SCALE
            h = dna["rows"].get(str(r), 15.0) * H_SCALE
            
            val = cell_values.get((r, c), "")
            is_label = val != ""
            klass = "label" if is_label else "input-box"
            field_id = f"cell_{letter}{r}"
            content = val if is_label else f'<input type="text" id="{field_id}">'
            
            html.append(f'  <div class="box {klass}" style="left: {x}px; top: {y}px; width: {w}px; height: {h}px;">{content}</div>')

    html.append("""
    </div>
    <script>
        const patientId = new URLSearchParams(window.location.search).get('id');
        async function loadData() {
            if (!patientId) return;
            const res = await fetch(`/api/plan?id=${patientId}`);
            const data = await res.json();
            const plan = data.plan || {};
            
            // Map common fields to cell IDs (Heuristic fallback)
            const fieldMap = {
                "patient_name": "cell_B4",
                "disease_name": "cell_B5",
            };

            document.querySelectorAll('input, textarea').forEach(el => {
                const dbKey = el.id.replace('cell_', '');
                if (plan[el.id]) {
                    el.value = plan[el.id];
                }
            });
        }

        async function saveData() {
            const payload = { patient_id: patientId, plan_date: new Date().toISOString().split('T')[0] };
            document.querySelectorAll('input, textarea').forEach(el => {
                if (el.id) payload[el.id] = el.value;
            });
            const res = await fetch('/api/save', { method: 'POST', body: JSON.stringify(payload) });
            if (res.ok) alert("保存完了。構造は1ミリの狂いもなく固定されています。");
        }
        window.onload = loadData;
    </script>
</body>
</html>
    """)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))
    print(f"Generated HARD-LOCKED clinical editor: {output_path}")

if __name__ == "__main__":
    generate_hard_locked_html("excel_dna.json", "templates/hard_locked_replica.html")
