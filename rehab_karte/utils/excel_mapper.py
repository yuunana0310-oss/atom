import openpyxl
import os
import shutil
import json
from datetime import datetime
from database.db import get_connection

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "master_template.xlsx")

# THE EXHAUSTIVE MAP: DB Field -> Excel Cell Address
EXCEL_MAP = {
    # Header
    "patient_name": "B4",
    "disease_name": "B5",
    "onset_date": "AB5",
    "surgery_date": "AB6",
    "sex": "K4",
    "age": "N4",
    
    # Section 1: Functions
    "impairment_consciousness": "E10",
    "impairment_respiratory": "Q10",
    "impairment_circulate": "E11",
    "impairment_rom": "Q11",
    "impairment_muscle": "E12",
    "impairment_paralysis": "Q12",
    "impairment_sensation": "E13",
    "impairment_higher_brain": "Q13",
    
    # Section 4: Goals/Plan (Page 2)
    "rehab_history": "B45",
    "goal_short": "B50",
    "goal_long": "B55",
    "plan_pt": "B60",
    "plan_ot": "B65",
    "plan_st": "B70",
    
    # Staff / Signatures
    "staff_physician": "C83",
    "therapist_pt": "C85",
    "therapist_ot": "I85",
    "staff_ns": "U85",
    "sign_date_dr": "O87",
    "sign_date_pt": "U87",
}

# ADL Matrix (Full 17-item FIM/BI)
ADL_START_ROW = 34
ADL_ITEM_ORDER = [
    "食事", "整容", "清拭・入浴", "更衣(上身)", "更衣(下身)", "トイレ動作", 
    "排尿管理", "排便管理", "移乗(ﾍﾞｯﾄﾞ等)", "移乗(ﾄイレ)", "移動(歩行等)", "移動(階段)", 
    "理解", "表出", "社会的交流", "問題解決", "記憶"
]
# Columns: Initial=F, Prev=H, Current=J, Goal=L
ADL_COL_MAP = {"start": "F", "prev": "H", "now": "J", "goal": "L"}

def export_to_excel(patient_id, plan_data):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_filename = f"様式21号_{patient_id}_{timestamp}.xlsx"
    target_path = os.path.join(desktop, output_filename)
    
    shutil.copy(TEMPLATE_PATH, target_path)
    wb = openpyxl.load_workbook(target_path)
    ws = wb.active
    
    # 1. Fill Absolute Fields
    for db_key, cell in EXCEL_MAP.items():
        if db_key in plan_data and plan_data[db_key]:
            ws[cell] = str(plan_data[db_key])
            
    # 2. Fill ADL Matrix
    if "adl_evaluation_json" in plan_data and plan_data["adl_evaluation_json"]:
        try:
            adl = json.loads(plan_data["adl_evaluation_json"])
            for idx, item_name in enumerate(ADL_ITEM_ORDER):
                if item_name in adl:
                    row = ADL_START_ROW + idx
                    for key, col in ADL_COL_MAP.items():
                        if key in adl[item_name]:
                            ws[f"{col}{row}"] = adl[item_name][key]
        except Exception as e:
            print(f"ADL Mapping Error: {e}")
            
    wb.save(target_path)
    return target_path

if __name__ == "__main__":
    pass
