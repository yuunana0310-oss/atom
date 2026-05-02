import openpyxl
import json
import os

path = r"C:\Users\yuuna\Desktop\e21 (1).xlsx"

def extract_dna(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    dna = {
        "title": ws.title,
        "cols": {},
        "rows": {},
        "merges": [str(m) for m in ws.merged_cells.ranges],
        "cells": []
    }
    
    # Extract Col Widths (1-100)
    for i in range(1, 101):
        col_letter = openpyxl.utils.get_column_letter(i)
        width = ws.column_dimensions[col_letter].width
        dna["cols"][col_letter] = width if width is not None else 8.38 # default

    # Extract Row Heights (1-100)
    for r in range(1, 121):
        height = ws.row_dimensions[r].height
        dna["rows"][r] = height if height is not None else 15.0 # default
        
        # Capture labels
        row_data = []
        for c in range(1, 101):
            val = ws.cell(r, c).value
            if val:
                dna["cells"].append({"r": r, "c": c, "v": str(val)})

    with open("excel_dna.json", "w", encoding="utf-8") as f:
        json.dump(dna, f, indent=2, ensure_ascii=False)
    
    print(f"DNA Extraction complete. Merged count: {len(dna['merges'])}")

if __name__ == "__main__":
    extract_dna(path)
