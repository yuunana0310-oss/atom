import openpyxl
import os

path = r"C:\Users\yuuna\Desktop\e21 (1).xlsx"

def signature_map(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    print("--- PAGE 2 SIGNATURE SCAN ---")
    for r in range(80, 105):
        line = []
        for c in range(1, 40):
            val = ws.cell(r, c).value
            line.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {val}" if val else "")
        row_str = " | ".join([x for x in line if x])
        if row_str:
            print(row_str)

if __name__ == "__main__":
    signature_map(path)
